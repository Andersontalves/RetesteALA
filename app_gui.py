"""
Aplicativo com Interface Gráfica (GUI) para Processador FTTH
"""
import sys
import os
import threading
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk

warnings.filterwarnings('ignore')

class ProcessadorFTTHApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Processador de Sinal FTTH - Analisador de OLTCloud")
        self.root.geometry("750x550")
        self.root.configure(padx=20, pady=20)
        
        self.arquivo_selecionado = None
        self.caminho_salvar = None
        
        # Configurar ícone se existir
        if getattr(sys, 'frozen', False):
            base_dir = sys._MEIPASS
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            
        icone_path = os.path.join(base_dir, "alares.ico")
        if os.path.exists(icone_path):
            try:
                self.root.iconbitmap(icone_path)
            except:
                pass
        
        self.setup_ui()
        
    def setup_ui(self):
        # Título
        titulo = ttk.Label(self.root, text="Cruzamento Base & OLTCloud", font=("Segoe UI", 16, "bold"), background=self.root.cget('bg'))
        titulo.pack(pady=(0, 20))
        
        # Frame de Seleção de Arquivo
        frame_arquivo = ttk.LabelFrame(self.root, text=" 1. Selecione a Planilha (.xlsx) ", padding=15)
        frame_arquivo.pack(fill=tk.X, pady=(0, 20))
        
        self.lbl_arquivo = ttk.Label(frame_arquivo, text="Nenhum arquivo selecionado", foreground="gray")
        self.lbl_arquivo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        btn_selecionar = ttk.Button(frame_arquivo, text="Buscar Arquivo", command=self.selecionar_arquivo)
        btn_selecionar.pack(side=tk.RIGHT)
        
        # O que o script vai fazer
        lbl_info = ttk.Label(self.root, text="Regras: A Aba 1 deve ser CLIENTES. A Aba 2 deve ser OLTCLOUD.", font=("Segoe UI", 9, "italic"))
        lbl_info.pack(anchor=tk.W, pady=(0, 20))
        
        # Botão de Ação
        self.btn_processar = ttk.Button(self.root, text="▶ INICIAR PROCESSAMENTO", command=self.iniciar_processamento, state=tk.DISABLED)
        self.btn_processar.pack(fill=tk.X, pady=(0, 20), ipady=5)
        
        # Assinatura do Criador no rodapé primeiro (garante espaço)
        lbl_autor = tk.Label(self.root, text="Created by: Anderson Tadeu Alves", font=("Segoe UI", 10, "bold italic"), fg="#333333", bg=self.root.cget('bg'))
        lbl_autor.pack(side=tk.BOTTOM, pady=(5, 0))

        # Frame de Logs (expande no restante do espaço)
        frame_logs = ttk.LabelFrame(self.root, text=" Logs de Processamento ", padding=10)
        frame_logs.pack(fill=tk.BOTH, expand=True)
        
        self.txt_logs = scrolledtext.ScrolledText(frame_logs, height=15, bg="black", fg="lightgreen", font=("Consolas", 10))
        self.txt_logs.pack(fill=tk.BOTH, expand=True)
        self.log("Aguardando seleção de arquivo...")

    def log(self, mensagem):
        self.txt_logs.insert(tk.END, mensagem + "\n")
        self.txt_logs.see(tk.END)
        self.root.update_idletasks()

    def selecionar_arquivo(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione a planilha",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if arquivo:
            self.arquivo_selecionado = arquivo
            self.lbl_arquivo.config(text=os.path.basename(arquivo), foreground="blue")
            self.btn_processar.config(state=tk.NORMAL)
            self.log(f"Arquivo selecionado: {arquivo}")

    def iniciar_processamento(self):
        if not self.arquivo_selecionado:
            return
            
        # Pede onde salvar o novo arquivo ANTES de iniciar a thread
        arquivobase = os.path.basename(self.arquivo_selecionado)
        novo_nome = "PROCESSADO_" + arquivobase
        caminho_salvar = filedialog.asksaveasfilename(
            title="Salvar resultado como...",
            initialfile=novo_nome,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not caminho_salvar:
            self.log("Operação cancelada (nenhum local de salvamento escolhido).")
            return
            
        self.caminho_salvar = caminho_salvar
        self.btn_processar.config(state=tk.DISABLED, text="PROCESSANDO... AGUARDE")
        self.txt_logs.delete(1.0, tk.END)
        self.log(f"Iniciando. Arquivo será salvo em:\n{self.caminho_salvar}\n")
        
        # Executa em thread separada para não travar a tela
        thread = threading.Thread(target=self.processar_dados)
        thread.daemon = True
        thread.start()

    def processar_dados(self):
        try:
            arquivo = self.arquivo_selecionado
            self.log(f"Abrindo arquivo: {os.path.basename(arquivo)}")
            
            # 1. Verifica Abas
            self.log("Lendo estrutura de abas do arquivo...")
            xf = pd.ExcelFile(arquivo, engine='openpyxl')
            if len(xf.sheet_names) < 2:
                raise ValueError(f"O arquivo precisa ter pelo menos 2 abas. Encontradas: {xf.sheet_names}")
            
            ABA_PRINCIPAL = xf.sheet_names[0]
            ABA_OLTCLOUD  = xf.sheet_names[1]
            self.log(f"-> 1ª Aba identificada (Clientes): {ABA_PRINCIPAL}")
            self.log(f"-> 2ª Aba identificada (OLTCloud): {ABA_OLTCLOUD}")
            
            # 2. Leitura
            self.log(f"\nCarregando dados da aba {ABA_PRINCIPAL}...")
            df_bp = pd.read_excel(arquivo, sheet_name=ABA_PRINCIPAL, engine='openpyxl', dtype={'contrato': str})
            self.log(f"-> {len(df_bp)} linhas carregadas da base principal.")

            self.log(f"\nCarregando dados da aba {ABA_OLTCLOUD} (isso pode levar alguns segundos)...")
            df_olt = pd.read_excel(arquivo, sheet_name=ABA_OLTCLOUD, engine='openpyxl', dtype={'External Contract ID': str})
            self.log(f"-> {len(df_olt)} linhas carregadas da OLTCloud.")
            
            # 3. Preparação das chaves
            self.log("\nPreparando dados e normalizando colunas...")
            df_bp['_key'] = df_bp['contrato'].astype(str).str.strip().str.split('.').str[0]
            df_olt['_key'] = df_olt['External Contract ID'].astype(str).str.strip().str.split('.').str[0]
            df_olt = df_olt[df_olt['_key'] != 'nan'].copy()
            
            df_olt = df_olt.rename(columns={'RX ONU': 'RX_ONU', 'RX OLT': 'RX_OLT'})
            df_olt['RX_ONU'] = pd.to_numeric(df_olt['RX_ONU'], errors='coerce')
            df_olt['RX_OLT'] = pd.to_numeric(df_olt['RX_OLT'], errors='coerce')

            # Normalização (milli-dBm)
            self.log("Analisando e convertendo sinais em milli-dBm...")
            SINAL_MIN = -50.0
            SINAL_MAX = -8.0
            
            def normaliza_sinal(v):
                if pd.isna(v):                        return None
                if SINAL_MIN <= v <= SINAL_MAX:       return round(v, 2)
                v_conv = v / 1000.0
                if SINAL_MIN <= v_conv <= SINAL_MAX:  return round(v_conv, 2)
                return None

            df_olt['RX_ONU'] = df_olt['RX_ONU'].apply(normaliza_sinal)
            df_olt['RX_OLT'] = df_olt['RX_OLT'].apply(normaliza_sinal)
            
            # 4. Agrupamento
            self.log("Cruzando informações e verificando o melhor sinal por cliente...")
            df_olt_valido = df_olt.dropna(subset=['RX_ONU', 'RX_OLT'], how='all')
            df_olt_agrupado = (
                df_olt_valido
                .sort_values('RX_ONU', ascending=False)
                .groupby('_key', as_index=False)
                .first()
            )
            
            # 5. Cruzamento
            df_merge = df_bp.merge(
                df_olt_agrupado[['_key', 'RX_ONU', 'RX_OLT', 'Status', 'OLT', 'SN ONU', 'Modelo']],
                on='_key',
                how='left'
            )
            
            df_merge['EXISTE NA OLTCLOUD'] = df_merge['RX_ONU'].notna().map({True: 'SIM', False: 'NÃO'})
            
            def status_rx(v, min_v, max_v):
                if pd.isna(v): return 'SEM DADOS'
                if min_v <= v <= max_v: return 'BOM'
                if v > max_v: return 'SINAL ALTO'
                return 'RUIM'

            df_merge['STATUS RX ONU'] = df_merge['RX_ONU'].apply(lambda v: status_rx(v, -24.99, -10))
            df_merge['STATUS RX OLT'] = df_merge['RX_OLT'].apply(lambda v: status_rx(v, -26.99, -10))
            
            # 6. Montagem das abas
            self.log("\nPreparando aba RESULTADO e SOMENTE_BONS...")
            base_cols = list(df_bp.columns)
            extra_cols = ['EXISTE NA OLTCLOUD', 'RX_ONU', 'RX_OLT', 'STATUS RX ONU', 'STATUS RX OLT', 'Status', 'SN ONU', 'Modelo']
            all_cols = base_cols + extra_cols
            
            df_resultado = df_merge[[c for c in all_cols if c in df_merge.columns]].copy()
            df_resultado = df_resultado.rename(columns={'RX_ONU': 'RX ONU', 'RX_OLT': 'RX OLT'})
            for col in ['RX ONU', 'RX OLT']:
                if col in df_resultado.columns: df_resultado[col] = df_resultado[col].round(2)
            df_resultado = df_resultado.drop(columns=['_key'], errors='ignore')

            mascara_bom = (
                (df_merge['RX_ONU'] >= -24.99) & (df_merge['RX_ONU'] <= -10) &
                (df_merge['RX_OLT'] >= -26.99) & (df_merge['RX_OLT'] <= -10)
            )
            df_bons = df_merge[mascara_bom].copy()
            df_somente_bons = df_bons[[c for c in all_cols if c in df_bons.columns]].copy()
            df_somente_bons = df_somente_bons.rename(columns={'RX_ONU': 'RX ONU', 'RX_OLT': 'RX OLT'})
            for col in ['RX ONU', 'RX OLT']:
                if col in df_somente_bons.columns: df_somente_bons[col] = df_somente_bons[col].round(2)
            df_somente_bons = df_somente_bons.drop(columns=['_key'], errors='ignore')
            
            self.log(f"-> RESULTADO finalizado: {len(df_resultado)} linhas geradas")
            self.log(f"-> SOMENTE_BONS finalizado: {len(df_somente_bons)} contratos com sinal adequado")
            
            # 7. Escrita no Excel
            self.log(f"\nAbrindo excel {os.path.basename(arquivo)} para reescrita (isso pode demorar).")
            wb = load_workbook(arquivo)
            
            def escreve_aba(wb, nome_aba, df, cor_cabecalho):
                self.log(f"Escrevendo aba plana: {nome_aba}...")
                if nome_aba in wb.sheetnames:
                    del wb[nome_aba]
                ws = wb.create_sheet(nome_aba)

                fill_cab  = PatternFill('solid', fgColor=cor_cabecalho)
                font_cab  = Font(bold=True, color='FFFFFF', size=10)
                align_cab = Alignment(horizontal='center', vertical='center', wrap_text=False)
                colunas_sinal = {'RX ONU', 'RX OLT'}
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill, cell.font, cell.alignment, cell.border = fill_cab, font_cab, align_cab, border

                fills = {
                    'BOM': PatternFill('solid', fgColor='C6EFCE'), 'RUIM': PatternFill('solid', fgColor='FFC7CE'),
                    'SEM DADOS': PatternFill('solid', fgColor='FFEB9C'), 'SINAL ALTO': PatternFill('solid', fgColor='DDEBF7'),
                    'SIM': PatternFill('solid', fgColor='C6EFCE'), 'NÃO': PatternFill('solid', fgColor='FFC7CE'),
                }
                status_cols = {'EXISTE NA OLTCLOUD', 'STATUS RX ONU', 'STATUS RX OLT'}

                for row_idx, row in enumerate(df.itertuples(index=False), 2):
                    for col_idx, (col_name, value) in enumerate(zip(df.columns, row), 1):
                        val = '' if (pd.isna(value) if not isinstance(value, str) else False) else value
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.border, cell.alignment = border, Alignment(vertical='center')
                        if col_name in colunas_sinal and val != '': cell.number_format = '0.00'
                        if col_name in status_cols and str(val) in fills: cell.fill = fills[str(val)]

                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col_name)), 15)
                
                ws.freeze_panes = 'A2'
                ws.auto_filter.ref = ws.dimensions

            escreve_aba(wb, 'RESULTADO', df_resultado, '1F4E79')
            escreve_aba(wb, 'SOMENTE_BONS', df_somente_bons, '375623')
            
            self.log(f"\nSalvando o NOVO arquivo em:\n{self.caminho_salvar}")
            self.log("Não feche o programa ainda!")
            wb.save(self.caminho_salvar)
            
            # 8. Sucesso Final
            self.log("\n" + "=" * 50)
            self.log("✅ PROCESSO CONCLUÍDO COM SUCESSO!")
            self.log(f"Foram geradas {len(df_somente_bons)} linhas com sinal bom.")
            self.log("Você já pode fechar este programa e abrir o Excel.")
            self.log("=" * 50)
            
            # Restaura botão e dá pop-up feliz!
            self.root.after(0, lambda: self.finalizar_com_sucesso())
            
        except Exception as e:
            msg_erro = traceback.format_exc()
            self.log("\n" + "❌" * 20)
            self.log("OCORREU UM ERRO DURANTE O PROCESSAMENTO:\n")
            self.log(msg_erro)
            self.log("❌" * 20)
            self.root.after(0, lambda: self.mostrar_erro_na_tela(str(e)))

    def finalizar_com_sucesso(self):
        self.btn_processar.config(state=tk.NORMAL, text="▶ INICIAR PROCESSAMENTO")
        messagebox.showinfo("Sucesso!", "Processamento concluído com sucesso! O arquivo foi atualizado.\n\nVocê já pode fechar este programa e abrir a planilha.")
        
    def mostrar_erro_na_tela(self, e_str):
        self.btn_processar.config(state=tk.NORMAL, text="▶ INICIAR PROCESSAMENTO")
        messagebox.showerror("Atenção - Erro Encontrado", f"Não foi possível concluir o processamento.\n\nMotivo:\n{e_str}\n\nO programa não vai ser fechado. Leia os detalhes na tela preta de log para saber o que aconteceu.")


if __name__ == "__main__":
    # Configurar estilo DPI high awareness no Windows
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except:
        pass
        
    root = tk.Tk()
    
    # Estilizando abas e botões
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("TButton", padding=6, font=("Segoe UI", 10, "bold"))
    style.configure("TLabelframe.Label", font=("Segoe UI", 10, "bold"))
    
    app = ProcessadorFTTHApp(root)
    root.mainloop()
