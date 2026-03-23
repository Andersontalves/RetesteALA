"""
Refatoração do processar.py para ambiente web.
Recebe um stream (BytesIO) de um arquivo Excel em memória, processa, 
e retorna outro stream com o resultado.
"""

import pandas as pd
import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings('ignore')

SINAL_MIN = -50.0
SINAL_MAX = -8.0

def normaliza_sinal(v):
    if pd.isna(v):                        return None
    if SINAL_MIN <= v <= SINAL_MAX:       return round(v, 2)
    v_conv = v / 1000.0
    if SINAL_MIN <= v_conv <= SINAL_MAX:  return round(v_conv, 2)
    return None

def status_rx_onu(v):
    if pd.isna(v): return 'SEM DADOS'
    if -24.99 <= v <= -10: return 'BOM'
    if v > -10: return 'SINAL ALTO'
    return 'RUIM'

def status_rx_olt(v):
    if pd.isna(v): return 'SEM DADOS'
    if -26.99 <= v <= -10: return 'BOM'
    if v > -10: return 'SINAL ALTO'
    return 'RUIM'

def escreve_aba(wb, nome_aba, df, cor_cabecalho='1F4E79'):
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]
    ws = wb.create_sheet(nome_aba)

    fill_cab  = PatternFill('solid', fgColor=cor_cabecalho)
    font_cab  = Font(bold=True, color='FFFFFF', size=10)
    align_cab = Alignment(horizontal='center', vertical='center', wrap_text=False)
    colunas_sinal = {'RX ONU', 'RX OLT'}
    border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill_cab
        cell.font = font_cab
        cell.alignment = align_cab
        cell.border = border

    fills = {
        'BOM':        PatternFill('solid', fgColor='C6EFCE'),
        'RUIM':       PatternFill('solid', fgColor='FFC7CE'),
        'SEM DADOS':  PatternFill('solid', fgColor='FFEB9C'),
        'SINAL ALTO': PatternFill('solid', fgColor='DDEBF7'),
        'SIM':        PatternFill('solid', fgColor='C6EFCE'),
        'NÃO':        PatternFill('solid', fgColor='FFC7CE'),
    }

    status_cols = {'EXISTE NA OLTCLOUD', 'STATUS RX ONU', 'STATUS RX OLT'}

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, (col_name, value) in enumerate(zip(df.columns, row), 1):
            if pd.isna(value) if not isinstance(value, str) else False:
                value = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            if col_name in colunas_sinal and value != '':
                cell.number_format = '0.00'
            
            if col_name in status_cols and str(value) in fills:
                cell.fill = fills[str(value)]

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), 10)
        for row_idx in range(2, min(202, ws.max_row + 1)):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v:
                max_len = max(max_len, min(len(str(v)), 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

def process_excel_file(file_stream):
    try:
        file_stream.seek(0)
        xf = pd.ExcelFile(file_stream, engine='openpyxl')
        
        if len(xf.sheet_names) < 2:
            raise ValueError(f"O arquivo precisa ter pelo menos 2 abas. Encontradas: {xf.sheet_names}")
        if len(xf.sheet_names) < 3:
            raise ValueError("O arquivo precisa ter uma 3ª aba preenchida com os dados de O.S. já em reteste na coluna B.")

        ABA_PRINCIPAL = xf.sheet_names[0]
        ABA_OLTCLOUD  = xf.sheet_names[1]
        ABA_RETESTE   = xf.sheet_names[2]

        file_stream.seek(0)
        df_bp = pd.read_excel(file_stream, sheet_name=ABA_PRINCIPAL, engine='openpyxl', dtype={'contrato': str})
        
        file_stream.seek(0)
        df_olt = pd.read_excel(file_stream, sheet_name=ABA_OLTCLOUD, engine='openpyxl', dtype={'External Contract ID': str})
        
        file_stream.seek(0)
        df_reteste_tab = pd.read_excel(file_stream, sheet_name=ABA_RETESTE, engine='openpyxl', dtype=str, header=0)

        os_ja_em_reteste = set(
            df_reteste_tab.iloc[:, 1].dropna()
            .astype(str).str.strip().str.split('.').str[0]
        )
        os_ja_em_reteste.discard('nan')

        df_bp['_key'] = df_bp['contrato'].astype(str).str.strip().str.split('.').str[0]
        df_olt['_key'] = df_olt['External Contract ID'].astype(str).str.strip().str.split('.').str[0]

        df_olt = df_olt[df_olt['_key'] != 'nan'].copy()
        df_olt = df_olt.rename(columns={'RX ONU': 'RX_ONU', 'RX OLT': 'RX_OLT'})
        
        if 'RX_ONU' not in df_olt.columns or 'RX_OLT' not in df_olt.columns:
            raise ValueError("A 2ª aba (OLTCloud) precisa conter as colunas 'RX ONU' e 'RX OLT'.")

        df_olt['RX_ONU'] = pd.to_numeric(df_olt['RX_ONU'], errors='coerce')
        df_olt['RX_OLT'] = pd.to_numeric(df_olt['RX_OLT'], errors='coerce')

        df_olt['RX_ONU'] = df_olt['RX_ONU'].apply(normaliza_sinal)
        df_olt['RX_OLT'] = df_olt['RX_OLT'].apply(normaliza_sinal)

        df_olt_valido = df_olt.dropna(subset=['RX_ONU', 'RX_OLT'], how='all')
        df_olt_agrupado = (
            df_olt_valido
            .sort_values('RX_ONU', ascending=False)
            .groupby('_key', as_index=False)
            .first()
        )

        cols_olt = ['_key', 'RX_ONU', 'RX_OLT']
        for c in ['Status', 'OLT', 'SN ONU', 'Modelo']:
            if c in df_olt_agrupado.columns: cols_olt.append(c)

        df_merge = df_bp.merge(
            df_olt_agrupado[cols_olt],
            on='_key',
            how='left'
        )

        df_merge['EXISTE NA OLTCLOUD'] = df_merge['RX_ONU'].notna().map({True: 'SIM', False: 'NÃO'})
        df_merge['STATUS RX ONU'] = df_merge['RX_ONU'].apply(status_rx_onu)
        df_merge['STATUS RX OLT'] = df_merge['RX_OLT'].apply(status_rx_olt)

        colunas_resultado = (
            list(df_bp.columns) +
            ['EXISTE NA OLTCLOUD', 'RX_ONU', 'RX_OLT', 'STATUS RX ONU', 'STATUS RX OLT', 'Status', 'SN ONU', 'Modelo']
        )
        df_resultado = df_merge[[c for c in colunas_resultado if c in df_merge.columns]].copy()
        df_resultado = df_resultado.rename(columns={'RX_ONU': 'RX ONU', 'RX_OLT': 'RX OLT'})
        
        for col in ['RX ONU', 'RX OLT']:
            if col in df_resultado.columns:
                df_resultado[col] = df_resultado[col].round(2)
        df_resultado = df_resultado[[c for c in df_resultado.columns if c != '_key']]

        mascara_bom = (
            (df_merge['RX_ONU'] >= -24.99) & (df_merge['RX_ONU'] <= -10) &
            (df_merge['RX_OLT'] >= -26.99) & (df_merge['RX_OLT'] <= -10)
        )
        df_bons = df_merge[mascara_bom].copy()

        colunas_bons = (
            list(df_bp.columns) +
            ['EXISTE NA OLTCLOUD', 'RX_ONU', 'RX_OLT', 'STATUS RX ONU', 'STATUS RX OLT', 'Status', 'SN ONU', 'Modelo']
        )
        df_somente_bons = df_bons[[c for c in colunas_bons if c in df_bons.columns]].copy()
        df_somente_bons = df_somente_bons.rename(columns={'RX_ONU': 'RX ONU', 'RX_OLT': 'RX OLT'})
        for col in ['RX ONU', 'RX OLT']:
            if col in df_somente_bons.columns:
                df_somente_bons[col] = df_somente_bons[col].round(2)
        df_somente_bons = df_somente_bons[[c for c in df_somente_bons.columns if c != '_key']]

        agora   = datetime.datetime.now()
        periodo = 'MANHÃ' if agora.hour < 12 else 'TARDE'
        data_str = agora.strftime('%d/%m/%Y') + ' - ' + periodo

        chave_nos = df_somente_bons.get('n_os', pd.Series(dtype=str)).astype(str).str.strip().str.split('.').str[0]
        mascara_novos = ~chave_nos.isin(os_ja_em_reteste)
        df_novos = df_somente_bons[mascara_novos].copy()

        col_bairro = df_bp.columns[32] if len(df_bp.columns) >= 33 else 'bairro2'
        colunas_rp = ['n_os', 'contrato', 'nome_cliente', 'servico',
                      'RX ONU', 'RX OLT', 'cidade', col_bairro, 'nome_chassi']
        cols_disp = [c for c in colunas_rp if c in df_novos.columns]
        df_reteste_pronto = df_novos[cols_disp].copy()
        
        if col_bairro in df_reteste_pronto.columns:
            df_reteste_pronto = df_reteste_pronto.rename(columns={col_bairro: 'bairro'})
            
        df_reteste_pronto.insert(0, 'DATA / PERÍODO', data_str)

        file_stream.seek(0)
        wb = load_workbook(file_stream)
        
        escreve_aba(wb, 'RESULTADO',     df_resultado,    cor_cabecalho='1F4E79')
        escreve_aba(wb, 'SOMENTE_BONS',  df_somente_bons, cor_cabecalho='375623')
        escreve_aba(wb, 'RETESTE_PRONTO', df_reteste_pronto, cor_cabecalho='843C0C')

        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)
        
        return output_stream
        
    except ValueError as e:
        raise e
    except Exception as e:
        raise RuntimeError(f"Erro ao processar as planilhas: detalhe técnico - {str(e)}")
