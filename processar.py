"""
processar.py — Análise de Sinal FTTH
Cruzamento: 1ª aba (clientes) x 2ª aba (OLTCloud)
Gera:  RESULTADO  e  SOMENTE_BONS  no mesmo arquivo.

Como usar:
  1. Coloque este script na mesma pasta do arquivo Excel.
  2. Execute:  python processar.py
  3. O script encontra o .xlsx automaticamente e usa
     a 1ª aba como base de clientes e a 2ª como OLTCloud.
"""

import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import os

warnings.filterwarnings('ignore')

# ─────────────────────────────────────
# 0. AUTO-DETECTA O ARQUIVO XLSX
# ─────────────────────────────────────
pasta = os.path.dirname(os.path.abspath(__file__))
arquivos = [f for f in glob.glob(os.path.join(pasta, '*.xlsx'))
            if 'processar' not in os.path.basename(f).lower()]

if not arquivos:
    print("❌ Nenhum arquivo .xlsx encontrado na pasta do script.")
    print(f"   Pasta verificada: {pasta}")
    input("Pressione Enter para sair...")
    raise SystemExit(1)

if len(arquivos) > 1:
    print("⚠️  Mais de um arquivo .xlsx encontrado. Usando o primeiro:")
    for i, f in enumerate(arquivos, 1):
        print(f"   {i}. {os.path.basename(f)}")

ARQUIVO = arquivos[0]
print(f"📂 Arquivo: {os.path.basename(ARQUIVO)}")

# Detecta nomes das abas por posição
xf = pd.ExcelFile(ARQUIVO, engine='openpyxl')
if len(xf.sheet_names) < 2:
    print(f"❌ O arquivo precisa ter pelo menos 2 abas. Encontradas: {xf.sheet_names}")
    input("Pressione Enter para sair...")
    raise SystemExit(1)

ABA_PRINCIPAL = xf.sheet_names[0]
ABA_OLTCLOUD  = xf.sheet_names[1]
print(f"   1ª aba (clientes) : [{ABA_PRINCIPAL}]")
print(f"   2ª aba (OLTCloud)  : [{ABA_OLTCLOUD}]")
print()

# ─────────────────────────────────────
# 1. LEITURA DAS BASES
# ─────────────────────────────────────
print(f"📖 Lendo [{ABA_PRINCIPAL}]...")
df_bp = pd.read_excel(ARQUIVO, sheet_name=ABA_PRINCIPAL, engine='openpyxl', dtype={'contrato': str})
print(f"   ✅ {len(df_bp)} linhas | {len(df_bp.columns)} colunas")

print(f"📖 Lendo [{ABA_OLTCLOUD}] (pode demorar alguns segundos)...")
df_olt = pd.read_excel(ARQUIVO, sheet_name=ABA_OLTCLOUD, engine='openpyxl',
                       dtype={'External Contract ID': str})
print(f"   ✅ {len(df_olt)} linhas | {len(df_olt.columns)} colunas")


# ─────────────────────────────────────
# 2. PREPARA CHAVES DE JOIN
# ─────────────────────────────────────
# Garante que ambos os lados são string sem decimais
df_bp['_key'] = df_bp['contrato'].astype(str).str.strip().str.split('.').str[0]
df_olt['_key'] = df_olt['External Contract ID'].astype(str).str.strip().str.split('.').str[0]

# Remove linhas sem chave na OLTCLOUD
df_olt = df_olt[df_olt['_key'] != 'nan'].copy()

# Renomeia colunas de sinal para trabalhar com segurança
df_olt = df_olt.rename(columns={
    'RX ONU': 'RX_ONU',
    'RX OLT': 'RX_OLT',
})

# Converte sinais para numérico
df_olt['RX_ONU'] = pd.to_numeric(df_olt['RX_ONU'], errors='coerce')
df_olt['RX_OLT'] = pd.to_numeric(df_olt['RX_OLT'], errors='coerce')

# ─────────────────────────────────────
# NORMALIZAÇÃO DE SINAL
# Alguns registros estão em milli-dBm (ex: -22758 = -22.758 dBm).
# Regra: se |v| > 50 e v/1000 cai no range válido → converter.
# Valores como -99.99 (sentinel "sem leitura") → NaN.
# Range físico válido FTTH: -8 a -50 dBm
# ─────────────────────────────────────
SINAL_MIN = -50.0
SINAL_MAX =  -8.0

def normaliza_sinal(v):
    if pd.isna(v):                        return None  # sem dado
    if SINAL_MIN <= v <= SINAL_MAX:       return round(v, 2)  # já válido em dBm
    v_conv = v / 1000.0
    if SINAL_MIN <= v_conv <= SINAL_MAX:  return round(v_conv, 2)  # era milli-dBm
    return None  # inválido real (ex: -99.99, 9500)

df_olt['RX_ONU'] = df_olt['RX_ONU'].apply(normaliza_sinal)
df_olt['RX_OLT'] = df_olt['RX_OLT'].apply(normaliza_sinal)

millidbm = (df_olt['RX_ONU'].notna().sum() + df_olt['RX_OLT'].notna().sum())
print(f"   ✅ Sinais normalizados (milli-dBm convertidos, inválidos reais → vazio)")

# ─────────────────────────────────────
# 3. AGRUPA OLTCLOUD POR CONTRATO
#    (pega o registro com melhor sinal — RX ONU mais próximo de -8)
# ─────────────────────────────────────
print("🔗 Agrupando OLTCloud por contrato...")

# Filtra linhas com pelo menos um sinal válido
df_olt_valido = df_olt.dropna(subset=['RX_ONU', 'RX_OLT'], how='all')

# Para contratos com múltiplas ONUs, pega a de melhor sinal (RX ONU mais alto = menos negativo)
df_olt_agrupado = (
    df_olt_valido
    .sort_values('RX_ONU', ascending=False)   # melhor sinal primeiro
    .groupby('_key', as_index=False)
    .first()
)

# ─────────────────────────────────────
# 4. CRUZAMENTO
# ─────────────────────────────────────
print("🔎 Cruzando bases...")

df_merge = df_bp.merge(
    df_olt_agrupado[['_key', 'RX_ONU', 'RX_OLT', 'Status', 'OLT', 'SN ONU', 'Modelo']],
    on='_key',
    how='left'
)

# Existe na OLTCloud?
df_merge['EXISTE NA OLTCLOUD'] = df_merge['RX_ONU'].notna().map({True: 'SIM', False: 'NÃO'})

# Status de Sinal RX ONU
def status_rx_onu(v):
    if pd.isna(v): return 'SEM DADOS'
    if -24.99 <= v <= -10: return 'BOM'
    if v > -10: return 'SINAL ALTO'
    return 'RUIM'

# Status de Sinal RX OLT
def status_rx_olt(v):
    if pd.isna(v): return 'SEM DADOS'
    if -26.99 <= v <= -10: return 'BOM'
    if v > -10: return 'SINAL ALTO'
    return 'RUIM'

df_merge['STATUS RX ONU'] = df_merge['RX_ONU'].apply(status_rx_onu)
df_merge['STATUS RX OLT'] = df_merge['RX_OLT'].apply(status_rx_olt)

# ─────────────────────────────────────
# 5. MONTA ABA RESULTADO
#    Todos da BASE_PRINCIPAL + colunas de cruzamento
# ─────────────────────────────────────
colunas_resultado = (
    list(df_bp.columns) +
    ['EXISTE NA OLTCLOUD', 'RX_ONU', 'RX_OLT', 'STATUS RX ONU', 'STATUS RX OLT', 'Status', 'SN ONU', 'Modelo']
)
df_resultado = df_merge[[c for c in colunas_resultado if c in df_merge.columns]].copy()
df_resultado = df_resultado.rename(columns={'RX_ONU': 'RX ONU', 'RX_OLT': 'RX OLT'})
# Arredonda sinal para 2 casas decimais
for col in ['RX ONU', 'RX OLT']:
    if col in df_resultado.columns:
        df_resultado[col] = df_resultado[col].round(2)
# Remove coluna auxiliar se escorregou
df_resultado = df_resultado[[c for c in df_resultado.columns if c != '_key']]

# ─────────────────────────────────────
# 6. MONTA ABA SOMENTE_BONS
#    Só clientes com sinal dentro do critério em AMBOS os canais
# ─────────────────────────────────────
mascara_bom = (
    (df_merge['RX_ONU'] >= -24.99) & (df_merge['RX_ONU'] <= -10) &
    (df_merge['RX_OLT'] >= -26.99) & (df_merge['RX_OLT'] <= -10)
)
df_bons = df_merge[mascara_bom].copy()

# Todas as colunas da BASE_PRINCIPAL + sinal
colunas_bons = (
    list(df_bp.columns) +
    ['EXISTE NA OLTCLOUD', 'RX_ONU', 'RX_OLT', 'STATUS RX ONU', 'STATUS RX OLT', 'Status', 'SN ONU', 'Modelo']
)
df_somente_bons = df_bons[[c for c in colunas_bons if c in df_bons.columns]].copy()
df_somente_bons = df_somente_bons.rename(columns={'RX_ONU': 'RX ONU', 'RX_OLT': 'RX OLT'})
# Arredonda sinal para 2 casas decimais
for col in ['RX ONU', 'RX OLT']:
    if col in df_somente_bons.columns:
        df_somente_bons[col] = df_somente_bons[col].round(2)
df_somente_bons = df_somente_bons[[c for c in df_somente_bons.columns if c != '_key']]

print(f"   ✅ RESULTADO: {len(df_resultado)} linhas")
print(f"   ✅ SOMENTE_BONS: {len(df_somente_bons)} linhas com sinal BOM")

# ─────────────────────────────────────
# 7. ESCREVE DE VOLTA NA PLANILHA
# ─────────────────────────────────────
print("💾 Salvando na planilha...")

wb = load_workbook(ARQUIVO)

def escreve_aba(wb, nome_aba, df, cor_cabecalho='1F4E79'):
    """Reescreve uma aba com os dados do DataFrame, formatando o cabeçalho."""
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]
    ws = wb.create_sheet(nome_aba)

    fill_cab  = PatternFill('solid', fgColor=cor_cabecalho)
    font_cab  = Font(bold=True, color='FFFFFF', size=10)
    align_cab = Alignment(horizontal='center', vertical='center', wrap_text=False)
    # Colunas de sinal que recebem formato numérico 0.00
    colunas_sinal = {'RX ONU', 'RX OLT'}
    border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # Cabeçalho
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill_cab
        cell.font = font_cab
        cell.alignment = align_cab
        cell.border = border

    # Dados
    fills = {
        'BOM':        PatternFill('solid', fgColor='C6EFCE'),
        'RUIM':       PatternFill('solid', fgColor='FFC7CE'),
        'SEM DADOS':  PatternFill('solid', fgColor='FFEB9C'),
        'SINAL ALTO': PatternFill('solid', fgColor='DDEBF7'),
        'SIM':        PatternFill('solid', fgColor='C6EFCE'),
        'NÃO':        PatternFill('solid', fgColor='FFC7CE'),
    }

    status_cols = {'EXISTE NA OLTCLOUD', 'STATUS RX ONU', 'STATUS RX OLT'}

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, (col_name, value) in enumerate(zip(df.columns, row), 1):
            # Converte NaN para vazio
            if pd.isna(value) if not isinstance(value, str) else False:
                value = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            # Formato numérico 0.00 para colunas de sinal
            if col_name in colunas_sinal and value != '':
                cell.number_format = '0.00'
            # Colorir colunas de status
            if col_name in status_cols and str(value) in fills:
                cell.fill = fills[str(value)]

    # Ajustar largura das colunas automaticamente
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), 10)
        # Amostra de até 200 linhas
        for row_idx in range(2, min(202, ws.max_row + 1)):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v:
                max_len = max(max_len, min(len(str(v)), 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Congelar cabeçalho
    ws.freeze_panes = 'A2'

    # Auto-filtro
    ws.auto_filter.ref = ws.dimensions

    print(f"   ✅ Aba '{nome_aba}' escrita: {len(df)} linhas × {len(df.columns)} colunas")

escreve_aba(wb, 'RESULTADO',     df_resultado,    cor_cabecalho='1F4E79')
escreve_aba(wb, 'SOMENTE_BONS',  df_somente_bons, cor_cabecalho='375623')

wb.save(ARQUIVO)
print()
print("=" * 55)
print(f"  ✅ CONCLUÍDO!")
print(f"     RESULTADO    : {len(df_resultado)} linhas")
print(f"     SOMENTE_BONS : {len(df_somente_bons)} linhas (sinal BOM)")
print(f"     Arquivo salvo: {os.path.basename(ARQUIVO)}")
print("=" * 55)
